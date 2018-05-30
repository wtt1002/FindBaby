# -*- coding: UTF-8 -*-
# by Wtting

from urllib import request
import os
from openpyxl import Workbook  # 写入Excel表所用
from openpyxl import load_workbook  # 读取Excel表所用
if __name__ == '__main__':
    if os.path.exists('E:\\PycharmProjects\\FindBaby\\data.xlsx'):
        print('存在')
        wb = load_workbook('data.xlsx')
        ws = wb.active  # 获取当前正在操作的表对象
    else:
        print('新建')
        #  创建Excel表并写入数据
        wb = Workbook()  # 创建Excel对象
        ws = wb.active  # 获取当前正在操作的表对象
        # 往表中写入标题行,以列表形式写入
        ws.append(['寻亲类别', '寻亲编号', '姓名', '性别', '出生日期', '失踪时身高', '失踪时间', '失踪人所在地',
                   '失踪地点', '寻亲者特征描述', '其他资料', '注册时间', '跟进志愿者'])

    # 访问网址
    url = 'http://www.baobeihuijia.com/list.aspx?tid=1&sex=&photo=&page=18'
    # 代理IP
    proxy = {'http': '203.174.112.13:3128'}
    # 创建ProxyHandler
    proxy_support = request.ProxyHandler(proxy)
    # 创建Opener
    opener = request.build_opener(proxy_support)
    # 添加User-Agent
    opener.addheaders = [('User-Agent',
                          'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36')]
    # opener.addheaders =
    # [('User-Agent','Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; en) Presto/2.8.131 Version/11.11')]
    # 安装Opener
    request.install_opener(opener)
    # 使用自己安装好的Opener
    # response = request.urlopen(url)
    # 读取相应信息并解码
    # html = response.read().decode("utf-8")
    # 打印信息
    # print(html)
    # 关闭response
    # response.close()