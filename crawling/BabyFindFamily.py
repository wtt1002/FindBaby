# -*- coding: UTF-8 -*-
"""
@Time: 2018/5/30 12:17
@Author: TingTing W
"""
import os
import random
import time
import json

from openpyxl import Workbook  # 写入Excel表所用
from openpyxl import load_workbook  # 读取Excel表所用
from crawling import CrawlPage, ParsePage
# reload(sys)
# sys.setdefaultencoding('utf-8')

global ws  # 全局工作表对象


def write_to_file(content):
    with open('FailHref.txt', 'a+', encoding='utf-8') as f:
        f.write(json.dumps(content, ensure_ascii=False)+'\n')
        f.close()


def main(offset):
    # 第几页
    print('     页数: ' + str(offset))
    url = 'http://www.baobeihuijia.com/list.aspx?tid=2&sex=&photo=&page=' + str(offset)
    # 获取列表页html
    html_total = CrawlPage.getPage(url)
    # 如果获取页面失败
    if html_total is None:
        write_to_file(url)
        return "failed"
    # 获取页面成功，则再继续获取详情页
    items = ParsePage.parse_total_page(html_total)
    # 对于每一个id
    for item in items:
        # 获取详情页面
        url_detail = 'http://www.baobeihuijia.com/view.aspx?type=2&id=' + str(item)
        # 设置随机时间延迟爬取详情页
        random_sleep_time = random.uniform(0.1, 2.1)
        print('     延迟: ' + str(random_sleep_time) + 's' + '    页数:' + str(offset))
        time.sleep(random_sleep_time)
        # 获取详情页面
        html_detail = CrawlPage.getPage(url_detail)
        if html_detail is None:
            write_to_file(url_detail)
            continue
        # 解析详情页
        info_items = ParsePage.get_detail_info(html_detail)
        # print(info_items)
        # 写数据
        ws.append([info_items[0], info_items[1].split(">")[1].split("<")[0], info_items[2],
                   info_items[3], info_items[4], info_items[5], info_items[6], info_items[7], info_items[8],
                   info_items[9], info_items[10], info_items[11], info_items[12]])
        # 保存数据
        wb.save('data.xlsx')


if __name__ == '__main__':
    if os.path.exists('E:\\PycharmProjects\\FindBaby\\data.xlsx'):
        print('存在')
        wb = load_workbook('data.xlsx')
        ws = wb.active  # 获取当前正在操作的表对象
    else:
        print('新建')
        #  创建Excel表并写入数据
        wb = Workbook()  # 创建Excel对象
        ws = wb.active  # 获取当前正在操作的表对象
        # 往表中写入标题行,以列表形式写入
        ws.append(['寻亲类别', '寻亲编号', '姓名', '性别', '出生日期', '失踪时身高', '失踪时间', '失踪人所在地',
                   '失踪地点', '寻亲者特征描述', '其他资料', '注册时间', '跟进志愿者'])
        wb.save('data.xlsx')
    for i in range(346, 500):
        flag = main(i)
        if flag == "failed":
            continue
        wb.save('data.xlsx')
        print('###SUCCESS###第' + str(i) + '页\n')
# wb.close('data.xlsx')