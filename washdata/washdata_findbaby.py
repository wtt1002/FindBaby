# -*- coding: UTF-8 -*-
"""
@Time: 2018/6/2 14:56
@Author: TingTing W
"""
import openpyxl


def read07excel(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_by_name('sheet1')
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()


def write07excel(path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'address'

    value = [["名称", "价格", "出版社", "语言"],
             ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
             ["暗时间", "32.4", "人民邮电出版社", "中文"],
             ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
    for i in range(0, 4):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))

    wb.save(path)
    print("写入数据成功！")


def split_address(address):
    address_splits = address.split(u',')
    # for addr in address_splits:
    #     print(addr)
    ws_des.append(address_splits)
    des_book.save('demo.xlsx')


if __name__ == '__main__':
    src_path = 'E:\\test.xlsx'
    src_book = openpyxl.load_workbook(src_path)
    ws_src = src_book.active
    des_book = openpyxl.Workbook()
    ws_des = des_book.active
    ws_des.append(['省份', '城市', '详细'])
    des_book.save('demo.xlsx')
    for i in range(2, ws_src.max_row + 1):
        value = ws_src.cell(row=i, column=9).value
        split_address(value)
        print(value)