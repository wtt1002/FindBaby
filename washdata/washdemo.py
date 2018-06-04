# -*- coding: UTF-8 -*-
"""
@Time: 2018/6/2 14:56
@Author: TingTing W
"""
import openpyxl
import json

def read07excel(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_by_name('sheet1')
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()


def write07excel(path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'address'

    value = [["名称", "价格", "出版社", "语言"],
             ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
             ["暗时间", "32.4", "人民邮电出版社", "中文"],
             ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
    for i in range(0, 4):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))

    wb.save(path)
    print("写入数据成功！")


def write_to_file(content):
    with open('FailString.txt', 'a+', encoding='utf-8') as f:
        f.write(json.dumps(content, ensure_ascii=False)+'\n')
        f.close()


def split_double_address(address):
    index = 0
    try:
        address_splits = address.split(u',')
    except AttributeError:
        return None
    for addr in address_splits:
        index += 1
    for i in range(index, 4):
        address_splits.append('kong')
    return address_splits


if __name__ == '__main__':
    src_path = 'E:\\babyfindfamily.xlsx'
    src_book = openpyxl.load_workbook(src_path)
    ws_src = src_book.active
    des_book = openpyxl.Workbook()
    ws_des = des_book.active
    ws_des.append(['失踪人所在省份', '失踪人所在城市', '详细', '空白', '失踪省份', '失踪城市', '详细', '空白'])
    des_book.save('babyfindfamily_addr.xlsx')
    for i in range(2, ws_src.max_row + 1):
        # 现在的地址信息
        srcValue = ws_src.cell(row=i, column=8).value
        nowAddr = split_double_address(srcValue)
        if nowAddr is None:
            write_to_file(srcValue)
            continue
        # 失踪地的地址信息
        value = ws_src.cell(row=i, column=9).value
        preAddr = split_double_address(value)
        if preAddr is None:
            write_to_file(value)
            continue
        # 数据拼接
        for addr in preAddr:
            nowAddr.append(addr)
        print(nowAddr)
        # 写入excel
        ws_des.append(nowAddr)
        des_book.save('babyfindfamily_addr.xlsx')