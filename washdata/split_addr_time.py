# -*- coding: UTF-8 -*-
"""
@Time: 2018/6/14 21:17
@Author: TingTing W
"""
import openpyxl
import json


def write_to_file(content):
    with open('FailString.txt', 'a+', encoding='utf-8') as f:
        f.write(json.dumps(content, ensure_ascii=False)+'\n')
        f.close()


def str_to_int(str):
    return int(str)


def split_double_address(address):
    index = 0
    try:
        address_splits = address.split(u',')
    except AttributeError:
        return None
    for addr in address_splits:
        index += 1
    for i in range(index, 4):
        address_splits.append('kong')
    return address_splits


def split_double_time(address):
    index = 0
    try:
        time_splits = address.split(u',')
    except AttributeError:
        return None
    time_splits[0] = str_to_int(time_splits[0])
    time_splits[1] = str_to_int(time_splits[1])
    time_splits[2] = str_to_int(time_splits[2])
    return time_splits


if __name__ == '__main__':
    # 源数据集文件
    src_path = 'E:\\未清洗.xlsx'
    src_book = openpyxl.load_workbook(src_path)
    ws_src = src_book.active
    # 目的数据文件
    des_book = openpyxl.Workbook()
    ws_des = des_book.active
    # ws_des.append(['失踪人所在省份', '失踪人所在城市', '详细', '空白', '失踪省份', '失踪城市', '详细', '空白'])
    ws_des.append(['出生年份', '出生月份', '出生日', '失踪年份', '失踪月份', '失踪日', '失踪人所在省份', '失踪人所在城市', '详细', '空白', '失踪省份', '失踪城市', '详细', '空白'])
    des_book.save('demo.xlsx')
    for i in range(2, ws_src.max_row + 1):
        # 现在的地址信息
        birthInfo = ws_src.cell(row=i, column=5).value
        birthDay = split_double_time(birthInfo)
        if birthDay is None:
            write_to_file(birthInfo)
            continue
        # 失踪地的地址信息
        lostInfo = ws_src.cell(row=i, column=7).value
        lostDay = split_double_time(lostInfo)
        if lostDay is None:
            write_to_file(lostInfo)
            continue
        # 现在的地址信息
        now_addr_info = ws_src.cell(row=i, column=8).value
        nowaddr = split_double_address(now_addr_info)
        if nowaddr is None:
            write_to_file(now_addr_info)
            continue
        # 失踪地的地址信息
        pre_addr_info = ws_src.cell(row=i, column=9).value
        preAddr = split_double_address(pre_addr_info)
        if preAddr is None:
            write_to_file(pre_addr_info)
            continue
        # 数据拼接
        for time in lostDay:
            birthDay.append(time)
        # print(nowAddr)
        for addr in preAddr:
            birthDay.append(addr)
        for addr in nowaddr:
            birthDay.append(addr)
        # 写入excel
        ws_des.append(birthDay)
        des_book.save('demo.xlsx')

